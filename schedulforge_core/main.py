import os
import tempfile
import re
import openpyxl
from openpyxl.utils import get_column_letter
from fastapi import FastAPI, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse

app = FastAPI()

# Enable CORS only for your frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["https://manas-mahawar.github.io"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"]
)

# Regex for course codes like ABC123 L / ABC123 T / ABC123 P
course_pattern = re.compile(r"[A-Z]{2,4}\d{2,4}\s*[LTP]", re.IGNORECASE)

wb = None  # Global workbook variable


def save_temp_file(upload: UploadFile) -> str:
    suffix = os.path.splitext(upload.filename)[-1]
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp.write(upload.file.read())
        return tmp.name


@app.get("/")
def root():
    return {"message": "Unified timetable API running."}


@app.post("/list_sheets/")
async def list_sheets(file: UploadFile = File(...)):
    global wb
    temp_path = save_temp_file(file)
    wb = openpyxl.load_workbook(temp_path)

    seen = set()
    sheet_infos = []
    for name in wb.sheetnames:
        cleaned = name.strip().upper()
        if not cleaned.startswith("PG TIME") and cleaned not in seen:
            sheet_infos.append({"index": len(sheet_infos)+1, "name": cleaned})
            seen.add(cleaned)

    return {"sheets": sheet_infos}


@app.post("/list_tutorial_groups/")
async def list_tutorial_groups(
    file: UploadFile = File(...),
    sheet_choice: int = Form(...)
):
    global wb
    seen = set()
    sheet_map = [
        (idx+1, name.strip().upper(), name)
        for idx, name in enumerate(wb.sheetnames)
        if not name.strip().upper().startswith("PG TIME") and name.strip().upper() not in seen and not seen.add(name.strip().upper())
    ]
    selected = next((o_name for idx, _, o_name in sheet_map if idx == sheet_choice), None)
    if not selected:
        return JSONResponse(content={"error": "Invalid sheet_choice"}, status_code=400)

    sheet = wb[selected]
    name_upper = sheet.title.strip().upper()
    header_row, start_row = (6, 9) if name_upper == "1ST YEAR B" else (6, 8) if name_upper == "4TH YEAR B" else (5, 7) if name_upper == "4TH YEAR A" else (5, 8)

    groups = []
    for col in range(1, sheet.max_column + 1):
        val = sheet.cell(row=header_row, column=col).value
        if val:
            cleaned = str(val).strip().replace('\xa0', '').replace('\u200b', '')
            if cleaned.upper() != "DAY":
                groups.append(cleaned)

    return {"tutorial_groups": groups}


@app.post("/timetable/")
async def get_timetable(
    file: UploadFile = File(...),
    sheet_choice: int = Form(...),
    tutorial_group: str = Form(...)
):
    global wb
    seen = set()
    sheet_map = [
        (idx+1, name.strip().upper(), name)
        for idx, name in enumerate(wb.sheetnames)
        if not name.strip().upper().startswith("PG TIME") and name.strip().upper() not in seen and not seen.add(name.strip().upper())
    ]
    selected = next((o_name for idx, _, o_name in sheet_map if idx == sheet_choice), None)
    if not selected:
        return JSONResponse(content={"error": "Invalid sheet_choice"}, status_code=400)

    sheet = wb[selected]
    name_upper = sheet.title.strip().upper()
    header_row, start_row = (6, 9) if name_upper == "1ST YEAR B" else (6, 8) if name_upper == "4TH YEAR B" else (5, 7) if name_upper == "4TH YEAR A" else (5, 8)

    # Find correct column
    group_column = None
    for col in range(1, sheet.max_column + 1):
        val = sheet.cell(row=header_row, column=col).value
        if val and str(val).strip().replace('\xa0', '').replace('\u200b', '') == tutorial_group:
            group_column = col
            break
    if not group_column:
        return JSONResponse(content={"error": f"Tutorial group '{tutorial_group}' not found."}, status_code=404)

    time_slots = [
        "08:00 AM", "08:50 AM", "09:40 AM", "10:30 AM", "11:20 AM", "12:10 PM",
        "01:00 PM", "01:50 PM", "02:40 PM", "03:30 PM", "04:20 PM", "05:10 PM",
        "06:00 PM", "06:50 PM"
    ]
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
    row_to_day_time = {}
    for d, day in enumerate(days):
        for t, time in enumerate(time_slots):
            row_idx = start_row + d * 28 + t * 2
            row_to_day_time[row_idx] = (day, time)

    timetable = {day: {} for day in days}
    processed_merged = set()

    for row in range(start_row, start_row + 28 * len(days)):
        coord = f"{get_column_letter(group_column)}{row}"
        cell = sheet.cell(row=row, column=group_column)
        subject = None
        subject_row = row

        for merged in sheet.merged_cells.ranges:
            if coord in merged:
                top_left = (merged.min_row, merged.min_col)
                if top_left not in processed_merged:
                    subject = sheet.cell(*top_left).value
                    subject_row = merged.min_row
                    processed_merged.add(top_left)
                break
        else:
            subject = cell.value

        if not subject:
            continue

        subject_str = str(subject).strip().replace('\xa0', '').replace('\u200b', '').upper()
        if subject_str == tutorial_group.upper() or subject_str == "DAY":
            continue

        match = course_pattern.search(subject_str)
        if match:
            course_code = match.group(0).strip().upper()
        else:
            if subject_str in {"VK", "LAB", "LAB-2"} or len(subject_str) < 4:
                continue
            course_code = None

        if course_code and subject_row in row_to_day_time:
            day, time = row_to_day_time[subject_row]
            timetable[day][time] = course_code

    # Merge consecutive same subjects in same day
    for day in timetable:
        previous_subject = None
        count = 1
        merged_slots = {}

        for time in time_slots:
            current_subject = timetable[day].get(time)
            if current_subject == previous_subject and current_subject:
                count += 1
                merged_slots[time] = None
            else:
                if previous_subject and count > 1:
                    prev_time = time_slots[time_slots.index(time)-count]
                    timetable[day][prev_time] = f"{previous_subject} ({count} slots)"
                count = 1
            previous_subject = current_subject

        if previous_subject and count > 1:
            prev_time = time_slots[len(time_slots)-count]
            timetable[day][prev_time] = f"{previous_subject} ({count} slots)"

        for t in merged_slots:
            if t in timetable[day]:
                del timetable[day][t]

    return {
        "sheet_name": sheet.title,
        "tutorial_group": tutorial_group,
        "time_slots": time_slots,
        "timetable": timetable
    }
