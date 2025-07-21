import openpyxl
from openpyxl.utils import get_column_letter
import re

# Load workbook
wb = openpyxl.load_workbook("TIMETABLEJULY TODEC24.xlsx")

# Build list: (original_name, stripped_name), excluding PG TIME TABLE
sheet_infos = [
    (name, name.strip()) for name in wb.sheetnames
    if name.strip().upper() != "PG TIME TABLE"
]

print("Available sheets in workbook:")
for idx, (_, stripped_name) in enumerate(sheet_infos, start=1):
    print(f"{idx}. {stripped_name}")

# Ask user to choose a sheet
while True:
    try:
        choice = int(input("Enter the number corresponding to your batch sheet: "))
        if 1 <= choice <= len(sheet_infos):
            original_sheet_name, stripped_name = sheet_infos[choice - 1]
            sheet = wb[original_sheet_name]
            sheet_name_upper = stripped_name.upper()
            break
        else:
            print("Invalid choice. Please enter a number from the list.")
    except ValueError:
        print("Invalid input. Please enter a number.")

# Decide header row and start row based on selected sheet
if sheet_name_upper == "1ST YEAR A":
    header_row = 4
    start_row = 7
elif sheet_name_upper == "1ST YEAR B":
    header_row = 6
    start_row = 9
else:
    header_row = 5
    start_row = 8

# Ask for tutorial group
tutorial_group = input("Enter your tutorial group (e.g., 2O34): ").strip()

# Find the tutorial group column in header_row
group_column = None
for col in range(1, sheet.max_column + 1):
    val = sheet.cell(row=header_row, column=col).value
    if val and str(val).strip() == tutorial_group:
        group_column = col
        break

if not group_column:
    print(f" Tutorial group '{tutorial_group}' not found in row {header_row}.")
    exit()

# Time slots and weekdays
time_slots = [
    "08:00 AM", "08:50 AM", "09:40 AM", "10:30 AM", "11:20 AM", "12:10 PM",
    "01:00 PM", "01:50 PM", "02:40 PM", "03:30 PM", "04:20 PM", "05:10 PM",
    "06:00 PM", "06:50 PM"
]
day_list = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

# Regex pattern to extract course codes like 'ABC123 L'
course_pattern = re.compile(r"[A-Z]{3}\d{3}\s+[LPT]")

# Build row_to_day_time mapping
row_to_day_time = {}
for day_idx, day in enumerate(day_list):
    for slot_idx, time in enumerate(time_slots):
        row_idx = start_row + day_idx * 28 + slot_idx * 2
        row_to_day_time[row_idx] = (day, time)

# Prepare timetable dict
dict_timetable = {day: {"course": [], "time": []} for day in day_list}

# Track processed merged cells
processed_merged = set()

# Process timetable cells
for row in range(start_row, start_row + 28 * len(day_list)):
    cell = sheet.cell(row=row, column=group_column)
    coord = f"{get_column_letter(group_column)}{row}"

    # Handle merged cells
    subject_row = row
    for merged in sheet.merged_cells.ranges:
        if coord in merged:
            top_left = (merged.min_row, merged.min_col)
            if top_left not in processed_merged:
                subject = sheet.cell(*top_left).value
                subject_row = merged.min_row
                processed_merged.add(top_left)
                break
            else:
                subject = None
                break
    else:
        subject = cell.value

    if not subject:
        continue

    subject_str = str(subject).strip().upper()

    # Skip if header text or tutorial group name
    if subject_str == tutorial_group.upper() or subject_str == "DAY":
        continue

    # Find course code using regex
    match = course_pattern.search(subject_str)
    if not match:
        continue

    course_code = match.group(0)
    # Optional: remove space inside course code, e.g., 'UEC301 L' -> 'UEC301L'
    # course_code = course_code.replace(" ", "")

    # Only add if row maps to a valid day/time
    if subject_row in row_to_day_time:
        day, time_str = row_to_day_time[subject_row]
        dict_timetable[day]["course"].append(course_code)
        dict_timetable[day]["time"].append(time_str)

# Output result
print(f"\nTimetable for Tutorial Group: {tutorial_group}\n")
for day in dict_timetable:
    print(f"{day}:")
    print(f"  Course → {dict_timetable[day]['course']}")
    print(f"  Time   → {dict_timetable[day]['time']}")
    print()